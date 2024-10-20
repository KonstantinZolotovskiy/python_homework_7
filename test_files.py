import zipfile
from pypdf import PdfReader
from conftest import zip_path
from openpyxl import load_workbook


def test_pdf_file():
    with zipfile.ZipFile(zip_path) as file_zip:
        with file_zip.open('English.pdf') as pdf:
            reader = PdfReader(pdf)
            text = reader.pages[5].extract_text()
            assert 'Are you eating out every weekend?' in text


def test_xlsx_file():
    with zipfile.ZipFile(zip_path) as file_zip:
        with file_zip.open('Table.xlsx') as xlsx:
            workbook = load_workbook(xlsx)
            sheet = workbook.active
            assert sheet.cell(row=2, column=2).value == 'Petrov'
            assert sheet['A3'].value == 'Semen'


def test_csv_file():
    with zipfile.ZipFile(zip_path) as file_zip:
        with file_zip.open('csv.csv') as csv:
            content = csv.read()
            assert 'Ivanov' in str(content)
